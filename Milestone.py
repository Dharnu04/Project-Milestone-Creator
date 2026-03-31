import streamlit as st
from datetime import date, timedelta
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import uuid

# ═══════════════════════════════════════════════════════
# PAGE CONFIG
# ═══════════════════════════════════════════════════════
st.set_page_config(page_title="Shopify Milestone Creator", layout="wide")

# ═══════════════════════════════════════════════════════
# CSS
# ═══════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif;}
.main{background-color:#0f1117;}
/*.stApp{background:linear-gradient(135deg,#0f1117 0%,#1a1d2e 100%);}*/
.stApp{background:linear-gradient(135deg,#0f1117 0%,#000000 100%);}

.app-header{background:linear-gradient(155deg,#670C0C 0%,#C10801 30%,#F16001 60%,#D9C3AB 100%);border-radius:16px;padding:28px 36px;margin-bottom:28px;}
.app-header h1{color:white;font-size:2rem;font-weight:700;margin:0;letter-spacing:-0.5px;}
.app-header p{color:rgba(255,255,255,0.85);font-size:0.95rem;margin:4px 0 0;}

.sec-title{color:#96bf48;font-size:0.7rem;font-weight:700;text-transform:uppercase;letter-spacing:1.8px;margin:16px 0 8px;}

.sub-box{background:rgba(150,191,72,0.05);border-left:3px solid rgba(150,191,72,0.4);border-radius:0 8px 8px 0;padding:10px 14px;margin:4px 0 10px 2px;}
/* Hide Streamlit auto-generated anchor links */
.app-header h1 a,
.app-header a {
    display: none !important;
}

/* milestone ref list */
.m-ref-list{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.07);border-radius:10px;padding:14px 18px;margin-top:10px;}
.m-ref-row{display:flex;align-items:center;gap:10px;padding:5px 0;border-bottom:1px solid rgba(255,255,255,0.05);}
.m-ref-row:last-child{border-bottom:none;}
.m-ref-num{background:rgba(150,191,72,0.15);color:#96bf48;border-radius:5px;padding:2px 8px;font-size:0.72rem;font-weight:700;font-family:'DM Mono',monospace;min-width:28px;text-align:center;}
.m-ref-label{color:rgba(255,255,255,0.65);font-size:0.82rem;}

/* preview cards */
.milestone-header{background:linear-gradient(90deg,#96bf48,#5e8e3e);color:white;padding:10px 18px;border-radius:8px 8px 0 0;font-weight:600;font-size:0.92rem;display:flex;justify-content:space-between;align-items:center;}
.milestone-body{background:rgba(255,255,255,0.04);border:1px solid rgba(150,191,72,0.2);border-top:none;border-radius:0 0 8px 8px;padding:0;margin-bottom:12px;}
.task-row{padding:8px 18px;border-bottom:1px solid rgba(255,255,255,0.05);color:rgba(255,255,255,0.8);font-size:0.86rem;display:flex;align-items:center;gap:8px;}
.task-row:last-child{border-bottom:none;}
.task-dot{width:6px;height:6px;background:#96bf48;border-radius:50%;flex-shrink:0;}
.custom-task-row{padding:8px 18px;border-bottom:1px solid rgba(255,255,255,0.05);color:rgba(200,230,150,0.9);font-size:0.86rem;display:flex;align-items:center;gap:8px;}
.custom-task-dot{width:6px;height:6px;background:#f0c040;border-radius:50%;flex-shrink:0;}

.day-badge{background:rgba(150,191,72,0.15);color:#96bf48;border:1px solid rgba(150,191,72,0.3);border-radius:20px;padding:2px 10px;font-size:0.75rem;font-weight:600;font-family:'DM Mono',monospace;}
.date-badge{background:rgba(255,255,255,0.06);color:rgba(255,255,255,0.6);border-radius:6px;padding:2px 8px;font-size:0.75rem;font-family:'DM Mono',monospace;}

.summary-strip{background:rgba(150,191,72,0.08);border:1px solid rgba(150,191,72,0.25);border-radius:10px;padding:14px 22px;display:flex;gap:28px;margin-bottom:20px;flex-wrap:wrap;}
.summary-item{display:flex;flex-direction:column;}
.summary-label{color:rgba(255,255,255,0.45);font-size:0.68rem;text-transform:uppercase;letter-spacing:1px;font-weight:500;}
.summary-value{color:#96bf48;font-size:1.3rem;font-weight:700;font-family:'DM Mono',monospace;}

.info-chip{display:inline-block;background:rgba(255,200,50,0.1);border:1px solid rgba(255,200,50,0.3);color:rgba(255,200,50,0.9);border-radius:6px;padding:3px 10px;font-size:0.75rem;margin-top:4px;}
.warn-chip{display:inline-block;background:rgba(255,100,50,0.1);border:1px solid rgba(255,100,50,0.3);color:rgba(255,140,80,0.95);border-radius:6px;padding:3px 10px;font-size:0.75rem;margin-top:4px;}

div[data-testid="stCheckbox"] label{background:rgba(255,255,255,0.05);border:1px solid rgba(255,255,255,0.1);border-radius:8px;padding:8px 14px;transition:all 0.2s;cursor:pointer;}
div[data-testid="stCheckbox"] label:hover{border-color:#96bf48;background:rgba(150,191,72,0.08);}
div[data-testid="stNumberInput"] input{background:rgba(255,255,255,0.06)!important;border:1px solid rgba(255,255,255,0.12)!important;border-radius:8px!important;color:white!important;}
div[data-testid="stTextInput"] input{background:rgba(255,255,255,0.06)!important;border:1px solid rgba(255,255,255,0.12)!important;border-radius:8px!important;color:white!important;}
.stButton button{background:linear-gradient(90deg,#96bf48,#5e8e3e)!important;color:white!important;border:none!important;border-radius:8px!important;font-weight:600!important;padding:10px 28px!important;font-size:0.95rem!important;}
.stButton button:hover{opacity:0.9!important;transform:translateY(-1px)!important;}
.stDownloadButton button{background:rgba(255,255,255,0.07)!important;color:rgba(255,255,255,0.85)!important;border:1px solid rgba(255,255,255,0.15)!important;border-radius:8px!important;font-weight:500!important;}
.stDownloadButton button:hover{border-color:#96bf48!important;color:#96bf48!important;background:rgba(150,191,72,0.06)!important;}
label{color:rgba(255,255,255,0.75)!important;font-weight:500!important;}
hr{border-color:rgba(255,255,255,0.08)!important;}
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════
# DATE HELPERS
# ═══════════════════════════════════════════════════════

def is_working_day(d: date) -> bool:
    return d.weekday() < 5  # Mon=0 … Fri=4; Sat=5, Sun=6 skipped

def add_working_days(start: date, days: int) -> date:
    cur, added = start, 0
    while added < days:
        cur += timedelta(days=1)
        if is_working_day(cur):
            added += 1
    return cur

def count_working_days(start: date, end: date) -> int:
    count, cur = 0, start
    while cur <= end:
        if is_working_day(cur):
            count += 1
        cur += timedelta(days=1)
    return max(count, 1)

def next_working_day(d: date) -> date:
    d += timedelta(days=1)
    while not is_working_day(d):
        d += timedelta(days=1)
    return d

def fmt_date(d: date) -> str:
    return d.strftime("%d %b %Y")


# ═══════════════════════════════════════════════════════
# MILESTONE DEFINITIONS
# ═══════════════════════════════════════════════════════

BASE_MILESTONE_DEFS = [
    {
        "key": "M1", "label": "Foundation & Setup", "base_days": 7,
        # base_tasks intentionally empty — all M1 tasks injected dynamically
        # based on client type (retainer/new) and theme choices
        "base_tasks": [],
    },
    {
        "key": "M2", "label": "Design Phase", "base_days": 7,
        "base_tasks": [
            "Style Guide Preparation and Approval",
            "Website Homepage Layout (Desktop + Mobile)",
            "Banner and Image Creation (Desktop + Mobile)",
            "About Us Page Design",
            "Collection Page Design",
            "Product Page Design",
        ],
    },
    {
        "key": "M3", "label": "Development Phase", "base_days": 8,
        "base_tasks": [
            "Homepage Setup in Shopify (Desktop + Mobile)",
            "Product Page Setup in Shopify (Desktop + Mobile)",
            "About Us Page Setup",
            "Collection Page Setup in Shopify",
            "Development & Testing",
            "Bug Fixing",
        ],
    },
    {
        "key": "M4", "label": "Content & Configuration", "base_days": 5,
        "base_tasks": [
            "Header Menu Setup",
            "Footer Menu Setup",
            "Filter Setup",
            "Website Content Creation",
            "Website Content Upload",
            "Shopify Backend Settings",
        ],
    },
    {
        "key": "M5", "label": "Policies & SEO Basics", "base_days": 3,
        "base_tasks": [
            "Privacy Policy",
            "Shipping Policy",
            "Terms and Conditions",
            "Return and Refund Policy",
            "Meta Title and Description",
            "Website Proofreading",
        ],
    },
    # M6 = SEO — injected dynamically if toggled
    {
        "key": "M7", "label": "QA & Testing", "base_days": 4,
        "base_tasks": [
            "Website 1st Level Checklist",
            "Website 2nd Level Checklist",
            "Website First Cut Demonstration",
            "1st Revision + Correction",
        ],
    },
    {
        "key": "M8", "label": "Launch & Handover", "base_days": 2,
        "base_tasks": [
            "Website Final Submission",
            "Accounts Clearance",
            "Shopify Ownership Transfer to Client",
        ],
    },
]

SEO_TASKS = [
    "Keyword Research and Mapping",
    "On-Page SEO for Product Pages",
    "On-Page SEO for Collection Pages",
    "Meta Tags Optimisation (Title & Description)",
    "Image Alt Text and Compression",
    "URL Structure Review",
    "Schema Markup Setup",
    "Google Search Console Setup",
    "Sitemap Submission",
]



# ═══════════════════════════════════════════════════════
# DATE RECALCULATOR  (used by post-generation adjuster)
# ═══════════════════════════════════════════════════════

def recalculate_dates(plan: list[dict], start_date: date) -> list[dict]:
    """Recompute start/end dates for all milestones from start_date,
    preserving each milestone's days count."""
    cur = start_date
    for m in plan:
        m["start"] = cur
        m["end"]   = add_working_days(cur, m["days"] - 1)
        cur        = next_working_day(m["end"])
    return plan

# ═══════════════════════════════════════════════════════
# PLAN BUILDER
# ═══════════════════════════════════════════════════════

def build_plan(cfg: dict) -> list[dict]:
    milestones_raw = []

    for defn in BASE_MILESTONE_DEFS:
        key   = defn["key"]
        days  = defn["base_days"]
        tasks = list(defn["base_tasks"])

        # ── M1 ──────────────────────────────────
        if key == "M1":
            # Manual base days from UI; conditionals add on top
            days = cfg.get("m1_base_days", 5)
            client_type = cfg.get("client_type", "new")  # "new" | "retainer"

            if client_type == "new":
                tasks += [
                    "Master Sheet Completion by Client",
                    "Theme Research",
                    "Theme Purchase on Shopify",
                    "Shopify Store Creation",
                    "Shopify Access to Client",
                ]
            else:
                retainer_theme = cfg.get("retainer_theme", "same")  # "same" | "new"
                if retainer_theme == "same":
                    tasks += ["Theme Research & Documentation"]
                else:
                    tasks += ["Theme Research", "Theme Approval & Purchase"]

            # Image assets from client
            if cfg.get("image_assets_from_client"):
                tasks += ["Collect Image Assets from Client", "Review & Approve Image Assets"]
                days += cfg.get("image_assets_days", 1)

            # Product images
            src = cfg.get("product_images_source", "none")
            if src == "client":
                tasks += ["Product Images Collection from Client", "Product Image Resizing / Fine Tune"]
                days += cfg.get("product_images_days", 2)
            elif src == "creation":
                tasks += ["Product Images Creation"]
                days += cfg.get("product_images_days", 3)

            # Logo
            if cfg.get("include_logo"):
                tasks += ["Logo Creation Brainstorm with Client", "Logo Creation"]
                days += cfg.get("logo_days", 2)

        # ── M2 ──────────────────────────────────
        if key == "M2":
            # Figma → Shopify flow
            if cfg.get("design_approach") == "figma":
                tasks = ["Layout Design in Figma"] + tasks
                days += cfg.get("figma_days", 3)

            # Custom development design
            if cfg.get("custom_dev"):
                tasks += ["Custom Development Design"]
                days += cfg.get("custom_dev_design_days", 3)

        # ── M3 ──────────────────────────────────
        if key == "M3":
            # App integration (moved from M1)
            if cfg.get("include_app"):
                tasks += ["App Integration in Shopify for Product Fetching"]
                days += cfg.get("app_days", 8)

            # Custom development
            if cfg.get("custom_dev"):
                tasks += ["Custom Development"]
                days += cfg.get("custom_dev_days", 5)

        # ── M7 ──────────────────────────────────
        if key == "M7" and cfg.get("include_payment"):
            tasks += ["Payment Gateway Creation", "Payment Gateway Testing"]
            days += cfg.get("payment_days", 2)

        milestones_raw.append({"key": key, "label": defn["label"], "days": days, "tasks": tasks})

    # Inject SEO before M7
    if cfg.get("include_seo"):
        seo = {
            "key": "M6", "label": "SEO",
            "days": cfg.get("seo_days", 6),
            "tasks": list(SEO_TASKS),
        }
        idx = next((i for i, m in enumerate(milestones_raw) if m["key"] == "M7"), len(milestones_raw))
        milestones_raw.insert(idx, seo)

    # Inject custom add-on tasks into their target milestones
    custom_tasks = cfg.get("custom_addon_tasks", [])
    for ct in custom_tasks:
        target_key = ct["milestone_key"]
        for m in milestones_raw:
            if m["key"] == target_key:
                m["tasks"].append(f"{ct['task_name']} ⭐")
                m["days"] += ct["days"]
                break

    # Proportional scaling to total_days
    raw_total = sum(m["days"] for m in milestones_raw)
    scale = cfg["total_days"] / raw_total if raw_total > 0 else 1.0

    plan, cur_start = [], cfg["start_date"]
    for i, m in enumerate(milestones_raw):
        scaled = max(1, round(m["days"] * scale))
        end = add_working_days(cur_start, scaled - 1)
        plan.append({
            "number": f"Milestone {i + 1}",
            "key": m["key"], "label": m["label"],
            "days": scaled, "tasks": m["tasks"],
            "start": cur_start, "end": end,
        })
        cur_start = next_working_day(end)

    return plan


# ═══════════════════════════════════════════════════════
# PDF EXPORT
# ═══════════════════════════════════════════════════════
def generate_pdf(plan: list[dict], cfg: dict) -> bytes:
    _uid = uuid.uuid4().hex[:8]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=1.8*cm, leftMargin=1.8*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    SG = colors.HexColor("#000000") # Header row BG, table outer BOX border, milestone separator lines, title text
    SL = colors.HexColor("#000000") # Thick right border after milestone column (can differ from SG)
    DB = colors.HexColor("#000000") # Main table header row background (dark)
    RA = colors.HexColor("#f5f5f5") # Alternating milestone stripe (even milestones) + summary row BG
    WH = colors.white                 # Header text color, summary header text
    LG = colors.HexColor("#e8f5d0")   # light green stripe for custom ⭐ tasks

    def ps(n, **kw): return ParagraphStyle(n, **kw)

    story = []

    # ── Title ──────────────────────────────────────────────────────────────
    story.append(Paragraph(
        f"{cfg['project_name']} — Milestone Plan",
        ps("T", fontName="Helvetica-Bold", fontSize=14,
           textColor=colors.HexColor("#000000"), spaceAfter=4)))
    # story.append(st_tbl)
    story.append(Spacer(1, 18))

    # ── Summary strip ──────────────────────────────────────────────────────
    sd = [["Project", "Client", "Platform", "Total Days", "Start", "End"],
          [cfg["project_name"], cfg.get("client_name", "—") or "—", "Shopify",
           str(cfg["total_days"]), fmt_date(cfg["start_date"]), fmt_date(plan[-1]["end"])]]
    st_tbl = Table(sd, colWidths=[3.8*cm, 3*cm, 2.2*cm, 2*cm, 2.6*cm, 2.6*cm])
    st_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), SG), ("TEXTCOLOR", (0, 0), (-1, 0), WH),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("FONTNAME",   (0, 1), (-1, 1), "Helvetica"),
        ("BACKGROUND", (0, 1), (-1, 1), RA),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#2c2c2c")),#--->
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(st_tbl)
    story.append(Spacer(1, 18))

    # ── Main table — one row per task, with row-spanning ──────────────────
    #
    # Column layout:
    #   0  Milestone name+number   3.8 cm   ← spans all task rows
    #   1  Task                    8.8 cm
    #   2  Days allocated          1.8 cm   ← spans
    #   3  Start date              2.5 cm   ← spans
    #   4  End date                2.5 cm   ← spans
    #
    COL_WIDTHS = [3.8*cm, 8.8*cm, 1.8*cm, 2.5*cm, 2.5*cm]

    # Header row
    tdata = [[
        Paragraph("<b>Milestones</b>",
                  ps(f"h0_{_uid}", fontName="Helvetica-Bold", fontSize=9,
                     textColor=WH, alignment=TA_CENTER)),
        Paragraph("<b>Tasks / Activity</b>",
                  ps(f"h1_{_uid}", fontName="Helvetica-Bold", fontSize=9,
                     textColor=WH, alignment=TA_LEFT)),
        Paragraph("<b>Days\nAllocated</b>",
                  ps(f"h2_{_uid}", fontName="Helvetica-Bold", fontSize=9,
                     textColor=WH, alignment=TA_CENTER)),
        Paragraph("<b>Start Date</b>",
                  ps(f"h3_{_uid}", fontName="Helvetica-Bold", fontSize=9,
                     textColor=WH, alignment=TA_CENTER)),
        Paragraph("<b>End Date</b>",
                  ps(f"h4_{_uid}", fontName="Helvetica-Bold", fontSize=9,
                     textColor=WH, alignment=TA_CENTER)),
    ]]

    span_cmds   = []   # SPAN TableStyle commands built as we go
    row_styles  = []   # background / font commands per row
    current_row = 1    # row 0 = header

    for m_idx, m in enumerate(plan):
        tasks      = m["tasks"] if m["tasks"] else ["—"]
        n_tasks    = len(tasks)
        first_row  = current_row
        last_row   = current_row + n_tasks - 1

        # Alternating milestone background (light green / white)
        ms_bg = RA if m_idx % 2 == 0 else WH

        for t_idx, task in enumerate(tasks):
            is_custom = task.endswith("⭐")
            task_bg   = colors.HexColor("#f0f8e0") if is_custom else ms_bg

            if t_idx == 0:
                # First task row — put milestone info in cols 0, 2, 3, 4
                row = [
                    Paragraph(
                        f"<b>{m['number']}</b><br/>"
                        f"<font size='7.5'>{m['label']}</font>",
                        ps(f"ms_{_uid}_{m_idx}",
                           fontName="Helvetica-Bold", fontSize=9, leading=13,
                           alignment=TA_CENTER)),
                    Paragraph(
                        # f"• {task}",
                        task,
                        ps(f"tk_{_uid}_{m_idx}_{t_idx}",
                           fontName="Helvetica", fontSize=8, leading=12)),
                    Paragraph(
                        f"<b>{m['days']}</b>",
                        ps(f"dy_{_uid}_{m_idx}",
                           fontName="Helvetica", fontSize=8,
                           alignment=TA_CENTER)),
                    Paragraph(
                        fmt_date(m["start"]),
                        ps(f"d1_{_uid}_{m_idx}",
                           fontName="Helvetica", fontSize=8,
                           alignment=TA_CENTER)),
                    Paragraph(
                        fmt_date(m["end"]),
                        ps(f"d2_{_uid}_{m_idx}",
                           fontName="Helvetica", fontSize=8,
                           alignment=TA_CENTER)),
                ]
            else:
                # Subsequent task rows — cols 0, 2, 3, 4 will be spanned (empty)
                row = [
                    "",
                    Paragraph(
                        # f"• {task}",
                        task,
                        ps(f"tk_{_uid}_{m_idx}_{t_idx}",
                           fontName="Helvetica", fontSize=8, leading=12)),
                    "", "", "",
                ]

            tdata.append(row)

            # Per-cell background for task column
            row_styles.append(("BACKGROUND", (1, current_row), (1, current_row), task_bg))
            # Background for spanned columns (milestone, days, start, end)
            for col in [0, 2, 3, 4]:
                row_styles.append(("BACKGROUND", (col, current_row), (col, current_row), ms_bg))

            current_row += 1

        # Register SPAN commands for cols 0, 2, 3, 4 across this milestone's rows
        if n_tasks > 1:
            for col in [0, 2, 3, 4]:
                span_cmds.append(("SPAN", (col, first_row), (col, last_row)))

        # Thick horizontal separator between milestones
        row_styles.append(
            ("LINEBELOW", (0, last_row), (-1, last_row), 1, SG)
        )

    # Build TableStyle
    base_style = [
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), DB),
        ("TEXTCOLOR",  (0, 0), (-1, 0), WH),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        # Global
        ("GRID",     (0, 0), (-1, -1), 0.3, colors.HexColor("#262626")), #----> 
        ("BOX",      (0, 0), (-1, -1), 1,   SG),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 7),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 7),
        ("VALIGN",    (0, 0), (-1, -1),  "TOP"),
        # Milestone / Days / Dates columns — center + middle-align
        ("ALIGN",  (0, 1), (0, -1), "CENTER"),
        ("ALIGN",  (2, 1), (4, -1), "CENTER"),
        ("VALIGN", (0, 1), (0, -1), "MIDDLE"),
        ("VALIGN", (2, 1), (4, -1), "MIDDLE"),
        # Thick right border after milestone column
        ("LINEAFTER", (0, 0), (0, -1), 1, SL),
    ]

    mt = Table(tdata, colWidths=COL_WIDTHS, repeatRows=1)
    mt.setStyle(TableStyle(base_style + row_styles + span_cmds))

    story.append(mt)
    story.append(Spacer(1, 14))
    story.append(Paragraph(
        f"Total Working Days: {cfg['total_days']} (can be extended by 4–5 days as buffer). "
        "All Design & Development Task will include a client review & Approval.",
        ps("ft", fontName="Helvetica-Oblique", fontSize=8,
           textColor=colors.HexColor("#888"), alignment=TA_CENTER)))

    doc.build(story)
    buf.seek(0)
    return buf.read()

# ═══════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════

def generate_excel(plan: list[dict], cfg: dict) -> bytes:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Milestone Plan"
    GD,GL,GP,WH,DK = "5E8E3E","96BF48","EBF5DF","FFFFFF","0F1117"

    def cs(cell, bold=False, bg=None, fg="000000", size=10, align="left", wrap=True, border=True):
        cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap, indent=1)
        if border:
            th = Side(style="thin", color="CCCCCC")
            cell.border = Border(left=th, right=th, top=th, bottom=th)

    ws.merge_cells("A1:F1"); ws["A1"] = f"{cfg['project_name']} — Shopify Milestone Plan"
    cs(ws["A1"], bold=True, bg=GD, fg=WH, size=14, align="center", border=False)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = (f"Platform: Shopify   |   Client: {cfg.get('client_name','—') or '—'}   |   "
                f"Working Days: {cfg['total_days']}   |   "
                f"Start: {fmt_date(cfg['start_date'])}   |   End: {fmt_date(plan[-1]['end'])}")
    cs(ws["A2"], bg=GL, fg=WH, size=9, align="center", border=False)
    ws.row_dimensions[2].height = 20; ws.append([])

    headers = ["Milestone","Phase","Tasks / Activity","Days Allocated","Start Date","End Date"]
    ws.append(headers); hr = ws.max_row
    for col, h in enumerate(headers, 1):
        cs(ws.cell(hr, col), bold=True, bg=DK, fg=WH, size=10, align="center")
    ws.row_dimensions[hr].height = 22

    for i, m in enumerate(plan):
        ws.append([m["number"], m["label"],
                   "\n".join(f"• {t}" for t in m["tasks"]),
                   m["days"], fmt_date(m["start"]), fmt_date(m["end"])])
        r = ws.max_row; bg = WH if i%2==0 else GP
        for col in range(1,7):
            cs(ws.cell(r,col), bg=bg, align="center" if col in [1,4,5,6] else "left", bold=(col in [1,4]))
        ws.row_dimensions[r].height = max(18, len(m["tasks"])*16)

    ws.append([]); ws.append(["","",f"Total Working Days: {cfg['total_days']} (+ 4–5 buffer days)","","",""])
    tr = ws.max_row; ws.merge_cells(f"C{tr}:F{tr}")
    cs(ws.cell(tr,3), bold=True, bg=GL, fg=WH, align="center", border=False)
    ws.cell(tr,1).fill = PatternFill("solid",fgColor=GP); ws.cell(tr,2).fill = PatternFill("solid",fgColor=GP)
    for i,w in enumerate([18,22,55,14,16,16],1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr+1}"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()


# ═══════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════

def days_input(label, key, default=2, min_val=1, max_val=60):
    return int(st.number_input(label, min_value=min_val, max_value=max_val,
                               value=default, step=1, key=key))

def section(title):
    st.markdown(f'<div class="sec-title">{title}</div>', unsafe_allow_html=True)

def sub_open():
    st.markdown('<div class="sub-box">', unsafe_allow_html=True)

def sub_close():
    st.markdown('</div>', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════
# SESSION STATE INIT
# ═══════════════════════════════════════════════════════

if "plan" not in st.session_state:
    st.session_state.plan = None
    st.session_state.config = None
    st.session_state.toast_msg = None
if "custom_tasks" not in st.session_state:
    st.session_state.custom_tasks = []   # list of {task_name, milestone_key, days}


# ═══════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════

st.markdown("""
<div class="app-header">
    <h1>Dharnu's Milestone Creator</h1>
    <p>Generate professional project milestone plans.</p>
</div>
""", unsafe_allow_html=True)

col_left, col_right = st.columns([1, 1.65], gap="large")


# ═══════════════════════════════════════════════════════
# LEFT PANEL
# ═══════════════════════════════════════════════════════

with col_left:

    # ── Project Details ─────────────────────────────────
    section("📋 Project Details")
    project_name = st.text_input("Project Name", placeholder="")
    client_name  = st.text_input("Client Name",  placeholder="QD/Zoketo")

    # ── Timeline ────────────────────────────────────────
    section("📅 Timeline")
    date_mode = st.radio("Set timeline by:", ["Start Date + Working Days", "Start Date + End Date"], horizontal=True)

    if date_mode == "Start Date + Working Days":
        dc1, dc2 = st.columns(2)
        with dc1: start_date = st.date_input("Start Date", value=date.today(), key="sd_wd")
        with dc2: total_days = days_input("Working Days", "wd_cnt", default=50, min_val=10, max_val=200)
        est_end = add_working_days(start_date, total_days - 1)
        st.markdown(f'<div class="info-chip">📆 Estimated end: <b>{fmt_date(est_end)}</b></div>', unsafe_allow_html=True)
    else:
        dc1, dc2 = st.columns(2)
        with dc1: start_date = st.date_input("Start Date", value=date.today(), key="sd_ed")
        with dc2: end_date   = st.date_input("End Date", value=date.today()+timedelta(weeks=8), key="ed_ed")
        if end_date <= start_date:
            st.error("End date must be after start date.")
            total_days = 1
        else:
            total_days = count_working_days(start_date, end_date)
            st.markdown(f'<div class="info-chip">🗓️ Calculated working days: <b>{total_days}</b></div>', unsafe_allow_html=True)

    st.markdown("---")

    # ── Client Type ─────────────────────────────────────
    section("👤 Client Type")
    client_type_radio = st.radio(
        "Is this a new or retainer client?",
        ["New Client", "Retainer Client"],
        key="client_type_rad", horizontal=True,
    )
    client_type = "new" if client_type_radio == "New Client" else "retainer"

    retainer_theme = None
    if client_type == "new":
        st.markdown(
            '<div class="info-chip">🆕 Full onboarding — Master Sheet, Theme Research & Purchase, '
            'Store Creation, Client Access included in M1</div>',
            unsafe_allow_html=True,
        )
    else:
        sub_open()
        retainer_theme_radio = st.radio(
            "Theme for this project:",
            ["Same Theme", "New Theme"],
            key="ret_theme_rad", horizontal=True,
        )
        retainer_theme = "same" if retainer_theme_radio == "Same Theme" else "new"
        if retainer_theme == "same":
            st.markdown(
                '<div class="info-chip">🔄 Theme Research & Documentation added to M1</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div class="info-chip">🆕 Theme Research + Theme Approval & Purchase added to M1</div>',
                unsafe_allow_html=True,
            )
        sub_close()

    # ── M1 Manual Base Days ──────────────────────────────
    section("📐 M1 Base Days")
    m1_base_days = days_input(
        "Foundation & Setup base days (extras add on top)",
        "m1_base_d", default=5, min_val=1, max_val=30,
    )
    st.markdown(
        f'<div class="info-chip">📌 M1 base = <b>{m1_base_days}d</b> + any enabled extras (logo, images etc.)</div>',
        unsafe_allow_html=True,
    )

    # ── Design Approach ─────────────────────────────────
    section("🖥️ Design Approach")
    design_approach = st.radio("Design workflow:", ["Direct Setup in Shopify", "Figma → Shopify"], key="design_rad", horizontal=True)
    figma_days = 0
    if design_approach == "Figma → Shopify":
        sub_open()
        figma_days = days_input("Days for Layout Design in Figma", "fd", default=3)
        sub_close()

    # ── Custom Development ──────────────────────────────
    section("⚙️ Custom Development")
    custom_dev = st.checkbox("Include Custom Development?", value=False, key="cdev_tog")
    custom_dev_design_days, custom_dev_days = 0, 0
    if custom_dev:
        sub_open()
        c1, c2 = st.columns(2)
        with c1: custom_dev_design_days = days_input("Custom Dev Design days (M2)", "cdd", default=3)
        with c2: custom_dev_days        = days_input("Custom Development days (M3)", "cdd2", default=5)
        sub_close()

    st.markdown("---")

    # ── Image Assets ────────────────────────────────────
    section("🖼️ Image Assets")
    image_assets_from_client = st.checkbox("Image assets provided by client?", value=False, key="ia_tog")
    image_assets_days = 0
    if image_assets_from_client:
        sub_open()
        image_assets_days = days_input("Days to collect & review", "iad", default=1)
        sub_close()

    # ── Product Images ──────────────────────────────────
    section("📦 Product Images")
    prod_img_src = st.radio("Product images source:", ["None / Not applicable","From Client","Create In-house"], key="pis")
    product_images_days = 0
    if prod_img_src == "From Client":
        sub_open()
        product_images_days = days_input("Days for collection & resizing / fine-tune", "pid_c", default=2)
        sub_close()
    elif prod_img_src == "Create In-house":
        sub_open()
        product_images_days = days_input("Days for product image creation", "pid_cr", default=3)
        sub_close()

    st.markdown("---")

    # ── Feature Toggles ─────────────────────────────────
    section("⚙️ Feature Toggles")

    include_logo = st.checkbox("🎨 Logo Creation", value=True, key="logo_tog")
    logo_days = 0
    if include_logo:
        sub_open(); logo_days = days_input("Days for logo creation", "ld", default=2); sub_close()

    include_app = st.checkbox("🔗 App Integration", value=False, key="app_tog")
    app_days = 0
    if include_app:
        sub_open(); app_days = days_input("Days for app integration", "ad", default=8, max_val=60); sub_close()

    include_payment = st.checkbox("💳 Payment Gateway", value=True, key="pay_tog")
    payment_days = 0
    if include_payment:
        sub_open(); payment_days = days_input("Days for payment gateway", "pd", default=2); sub_close()

    include_seo = st.checkbox("🔍 SEO Setup", value=False, key="seo_tog")
    seo_days = 0
    if include_seo:
        sub_open(); seo_days = days_input("Days for SEO", "sd", default=6, max_val=30); sub_close()

    st.markdown("---")

    # ── Custom Add-on Tasks ─────────────────────────────
    section("➕ Add-on Requirements")

    # Build live milestone reference list (dynamic based on toggles)
    def get_milestone_ref_list():
        refs = [
            ("M1","Foundation & Setup"),
            ("M2","Design Phase"),
            ("M3","Development Phase"),
            ("M4","Content & Configuration"),
            ("M5","Policies & SEO Basics"),
        ]
        idx = 6
        if include_seo:
            refs.append(("M6","SEO"))
            idx = 7
        refs.append((f"M{idx}","QA & Testing"))
        refs.append((f"M{idx+1}","Launch & Handover"))
        return refs

    include_custom = st.checkbox("Add custom tasks?", value=False, key="custom_tog")

    if include_custom:
        milestone_refs = get_milestone_ref_list()

        # Milestone reference helper
        rows_html = "".join(
            f'<div class="m-ref-row"><span class="m-ref-num">{k}</span>'
            f'<span class="m-ref-label">{v}</span></div>'
            for k, v in milestone_refs
        )
        st.markdown(f'<div class="m-ref-list">{rows_html}</div>', unsafe_allow_html=True)

        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

        # Add task form
        with st.form("add_custom_task_form", clear_on_submit=True):
            fc1, fc2, fc3 = st.columns([2.5, 1.5, 0.8])
            with fc1:
                new_task_name = st.text_input("Task Name", placeholder="e.g. A+ Images Creation")
            with fc2:
                milestone_options = [f"{k} – {v}" for k, v in milestone_refs]
                new_task_milestone = st.selectbox("Add to Milestone", milestone_options)
            with fc3:
                new_task_days = st.number_input("Days", min_value=1, max_value=30, value=2, step=1)
            add_btn = st.form_submit_button("＋ Add Task", use_container_width=True)

        if add_btn and new_task_name.strip():
            selected_key = new_task_milestone.split("–")[0].strip()
            st.session_state.custom_tasks.append({
                "task_name":    new_task_name.strip(),
                "milestone_key": selected_key,
                "days":         int(new_task_days),
                "milestone_label": new_task_milestone,
            })

        # Show added tasks
        if st.session_state.custom_tasks:
            st.markdown("<div style='margin-top:8px;'>", unsafe_allow_html=True)
            for i, ct in enumerate(st.session_state.custom_tasks):
                cc1, cc2 = st.columns([4, 1])
                with cc1:
                    st.markdown(
                        f'<div style="background:rgba(150,191,72,0.07);border:1px solid rgba(150,191,72,0.2);'
                        f'border-radius:7px;padding:7px 12px;font-size:0.82rem;color:rgba(255,255,255,0.8);">'
                        f'⭐ <b>{ct["task_name"]}</b> → {ct["milestone_label"]} '
                        f'<span style="color:#96bf48;">({ct["days"]}d)</span></div>',
                        unsafe_allow_html=True
                    )
                with cc2:
                    if st.button("✕", key=f"del_ct_{i}", use_container_width=True):
                        st.session_state.custom_tasks.pop(i)
                        st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
    else:
        # Clear custom tasks if toggle turned off
        if st.session_state.custom_tasks:
            st.session_state.custom_tasks = []

    st.markdown("---")

    # ── Milestone Reference (always visible at bottom) ──
    section("📌 Milestone Reference")
    milestone_refs_display = get_milestone_ref_list() if include_custom else [
        ("M1","Foundation & Setup"), ("M2","Design Phase"), ("M3","Development Phase"),
        ("M4","Content & Configuration"), ("M5","Policies & SEO Basics"),
        *([("M6","SEO")] if include_seo else []),
        (f"M{'7' if include_seo else '6'}","QA & Testing"),
        (f"M{'8' if include_seo else '7'}","Launch & Handover"),
    ]
    rows_html = "".join(
        f'<div class="m-ref-row"><span class="m-ref-num">{k}</span>'
        f'<span class="m-ref-label">{v}</span></div>'
        for k, v in milestone_refs_display
    )
    st.markdown(f'<div class="m-ref-list">{rows_html}</div>', unsafe_allow_html=True)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
    generate_btn = st.button("⚡ Generate Milestone Plan", use_container_width=True)


# ═══════════════════════════════════════════════════════
# GENERATE
# ═══════════════════════════════════════════════════════

if generate_btn:
    if not project_name.strip():
        st.error("Please enter a Project Name.")
    elif total_days < 1:
        st.error("Timeline is invalid — check your dates.")
    else:
        src_map = {"None / Not applicable":"none","From Client":"client","Create In-house":"creation"}
        cfg = {
            "project_name":            project_name.strip(),
            "client_name":             client_name.strip(),
            "start_date":              start_date,
            "total_days":              total_days,
            # Client type
            "client_type":             client_type,
            "retainer_theme":          retainer_theme,
            "m1_base_days":            m1_base_days,
            # Design
            "design_approach":         "figma" if design_approach == "Figma → Shopify" else "direct",
            "figma_days":              figma_days,
            # Custom dev
            "custom_dev":              custom_dev,
            "custom_dev_design_days":  custom_dev_design_days,
            "custom_dev_days":         custom_dev_days,
            # Images
            "image_assets_from_client": image_assets_from_client,
            "image_assets_days":        image_assets_days,
            "product_images_source":    src_map[prod_img_src],
            "product_images_days":      product_images_days,
            # Toggles
            "include_logo":    include_logo,    "logo_days":     logo_days,
            "include_app":     include_app,     "app_days":      app_days,
            "include_payment": include_payment, "payment_days":  payment_days,
            "include_seo":     include_seo,     "seo_days":      seo_days,
            # Custom add-ons
            "custom_addon_tasks": list(st.session_state.custom_tasks),
        }
        already_existed = st.session_state.plan is not None 
        st.session_state.plan   = build_plan(cfg)
        st.session_state.config = cfg
        st.session_state.toast_msg = "updated" if already_existed else "generated"


# ═══════════════════════════════════════════════════════
# RIGHT PANEL — Preview
# ═══════════════════════════════════════════════════════

with col_right:
    if st.session_state.get("toast_msg") == "generated":
        st.toast("✅ Milestone Generated!", icon="🎉")
        st.session_state.toast_msg = None
    elif st.session_state.get("toast_msg") == "updated":
        st.toast("🔄 Milestone Updated!", icon="✏️")
        st.session_state.toast_msg = None
    if st.session_state.plan:
        plan = st.session_state.plan
        cfg  = st.session_state.config
        total_tasks = sum(len(m["tasks"]) for m in plan)
        proj_end    = plan[-1]["end"]

        st.markdown(f"""
        <div class="summary-strip">
            <div class="summary-item"><span class="summary-label">Milestones</span><span class="summary-value">{len(plan)}</span></div>
            <div class="summary-item"><span class="summary-label">Total Tasks</span><span class="summary-value">{total_tasks}</span></div>
            <div class="summary-item"><span class="summary-label">Working Days</span><span class="summary-value">{cfg['total_days']}</span></div>
            <div class="summary-item"><span class="summary-label">Project End</span><span class="summary-value" style="font-size:1rem;">{fmt_date(proj_end)}</span></div>
            <div class="summary-item"><span class="summary-label">Client Type</span><span class="summary-value" style="font-size:0.95rem;">{"🔄 Retainer" if cfg.get("client_type")=="retainer" else "🆕 New"}</span></div>
        </div>
        """, unsafe_allow_html=True)

        for m in plan:
            tasks_html = ""
            for t in m["tasks"]:
                is_custom = t.endswith("⭐")
                if is_custom:
                    tasks_html += f'<div class="custom-task-row"><div class="custom-task-dot"></div>{t}</div>'
                else:
                    tasks_html += f'<div class="task-row"><div class="task-dot"></div>{t}</div>'

            st.markdown(f"""
            <div>
                <div class="milestone-header">
                    <span>{m['number']} — {m['label']}</span>
                    <div style="display:flex;gap:8px;align-items:center;">
                        <span class="date-badge">{fmt_date(m['start'])} → {fmt_date(m['end'])}</span>
                        <span class="day-badge">{m['days']} days</span>
                    </div>
                </div>
                <div class="milestone-body">{tasks_html}</div>
            </div>
            """, unsafe_allow_html=True)

        # ── Day Adjuster ─────────────────────────────────
        st.markdown("---")
        st.markdown('<div class="sec-title">🎛️ Adjust Milestone Days</div>', unsafe_allow_html=True)
        st.markdown(
            '<div style="color:rgba(255,255,255,0.45);font-size:0.8rem;margin-bottom:10px;">' +
            'Select a milestone, set new day count, and apply. All subsequent dates cascade automatically.' +
            '</div>',
            unsafe_allow_html=True
        )

        adj_options = [f"{m['number']} — {m['label']}" for m in plan]
        adj_col1, adj_col2, adj_col3, adj_col4 = st.columns([2.5, 0.6, 0.9, 0.6])

        with adj_col1:
            adj_selected = st.selectbox("Milestone", adj_options, key="adj_ms", label_visibility="collapsed")
        
        adj_idx = adj_options.index(adj_selected)
        current_days = plan[adj_idx]["days"]

        with adj_col2:
            if st.button("−", key="adj_minus", use_container_width=True):
                if "adj_days_val" not in st.session_state:
                    st.session_state.adj_days_val = current_days
                st.session_state.adj_days_val = max(1, st.session_state.adj_days_val - 1)

        with adj_col3:
            if "adj_days_val" not in st.session_state or st.session_state.get("_last_adj_idx") != adj_idx:
                st.session_state.adj_days_val = current_days
                st.session_state._last_adj_idx = adj_idx
            adj_days_val = st.number_input(
                "Days", min_value=1, max_value=120,
                value=st.session_state.adj_days_val,
                step=1, key="adj_days_input", label_visibility="collapsed"
            )
            st.session_state.adj_days_val = adj_days_val

        with adj_col4:
            if st.button("+", key="adj_plus", use_container_width=True):
                st.session_state.adj_days_val = st.session_state.adj_days_val + 1

        apply_col, _ = st.columns([1, 2])
        with apply_col:
            if st.button("✅ Apply Adjustment", key="adj_apply", use_container_width=True):
                new_days = st.session_state.adj_days_val
                st.session_state.plan[adj_idx]["days"] = new_days
                # Recalculate all dates from project start
                st.session_state.plan = recalculate_dates(
                    st.session_state.plan, st.session_state.config["start_date"]
                )
                # Update total days in config
                st.session_state.config["total_days"] = sum(m["days"] for m in st.session_state.plan)
                st.rerun()

        # ── Export ────────────────────────────────────────
        st.markdown("---")
        st.markdown('<div class="sec-title">📤 Export</div>', unsafe_allow_html=True)
        ecol1, ecol2 = st.columns(2)
        fname = cfg["project_name"].replace(" ", "_") or "project"
        with ecol1:
            st.download_button("📄 Download PDF",
                data=generate_pdf(plan, cfg),
                file_name=f"{fname}_milestone_plan.pdf",
                mime="application/pdf", use_container_width=True)
        with ecol2:
            st.download_button("📊 Download Excel",
                data=generate_excel(plan, cfg),
                file_name=f"{fname}_milestone_plan.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    else:
        st.markdown("""
        <div style="border:2px dashed rgba(150,191,72,0.25);border-radius:16px;padding:64px 32px;text-align:center;margin-top:20px;">
            <div style="font-size:3rem;margin-bottom:16px;"></div>
            <div style="color:rgba(255,255,255,0.5);font-size:1rem;line-height:1.7;">
                Fill in project details on the left<br>
                and click <strong style="color:#96bf48;">Generate Milestone Plan</strong><br>
                to preview your milestone breakdown here.
            </div>
        </div>
        """, unsafe_allow_html=True)
